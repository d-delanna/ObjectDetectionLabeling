import webbrowser, pyautogui
import string
from openpyxl import load_workbook
from pynput.keyboard import Key, Listener


class DataLabeling:
    def __init__(self, file_name, start_idx=2, machine='Windows'):
        print("Press \'Up\' for Yes, \'Down\' for No, "
              "\'Right\' for Unsure, and \'Left\' to undo."
              "\nTo exit, press \'Esc\'.\n")
        self.file_name, self.idx = file_name, start_idx
        self.workbook = load_workbook(filename=file_name)
        self.sheet = self.workbook.active
        self.col_idx = self.__get_columns()  # col idx of object_type, user_response, and url
        self.exiting = False
        key_dict = {'Windows': 'ctrl', 'Mac': 'command'}
        self.hot_key = key_dict[machine]
        self.__label_data()

    def __get_columns(self):
        col_names = [name.value for name in self.sheet['1']]
        col_idx = []
        for name in ['object', 'y/u/n', 'url']:
            idx = string.ascii_lowercase[col_names.index(name)].capitalize()
            col_idx.append(idx)
        return col_idx

    def __on_release(self, key):
        if key == Key.left:
            self.idx -= 1
            self.sheet.cell(row=self.idx, column=3).value = None
            print("Back/Undo")
        elif key in [Key.up, Key.down, Key.right]:
            input_dict = {Key.up: ('y', 'Yes'), Key.down: ('n', 'No'), Key.right: ('u', 'Pass/Unsure')}
            self.sheet.cell(row=self.idx, column=3).value = input_dict[key][0]
            self.idx += 1
            print(input_dict[key][1])
        elif key == Key.esc:
            self.exiting = True
            print("\nExiting...")
        else:
            print("Invalid Input")
        pyautogui.hotkey(self.hot_key, 'w')
        return False

    def __write_input(self):
        with Listener(on_release=self.__on_release) as listener:
            listener.join()

    def __label_data(self):
        while not self.exiting:
            print(self.idx, self.sheet[f'{self.col_idx[0]}{self.idx}'].value)
            if not self.sheet[f'{self.col_idx[1]}{self.idx}'].value:
                try:
                    url = self.sheet[f'{self.col_idx[2]}{self.idx}'].value
                    webbrowser.open(url, autoraise=True)
                    self.__write_input()
                except NameError:
                    print("\nNo URL found or data set completed.")
                    break
            else:
                print("Skipping... (cell already filled out)")
                self.idx += 1
            self.workbook.save(filename=self.file_name)


def main(file_name):
    DataLabeling(file_name)


if __name__ == '__main__':
    main(r'file\path\here')
